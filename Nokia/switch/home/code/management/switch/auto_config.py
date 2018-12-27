import commands
import os, time
import re, sys, paramiko
from openpyxl import Workbook
from openpyxl import load_workbook


def read_excel(IP, excel_path, sheet_name):
    command_list = []
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    for rx in ws.iter_rows(min_row=2, max_col=8, max_row=200):
        if not rx[0].value:
            break
        elif 'xx.xx.xx.xx' in rx[0].value:
            rx[0].value = rx[0].value.replace('xx.xx.xx.xx', IP)
        command_list.append(rx[0].value)
    try:
        if run_command(IP, command_list):
            return True
        else:
            return False
    except Exception, e:
        if telnet_command(IP, command_list, sheet_name):
            return True
        else:
            return False
    

def run_command(IP, command_list):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=IP, port=22, username='manager', password='nokia123', timeout=5)
    ssh_con = ssh.invoke_shell()
    ssh_con.send("\r\n")
    error = 0
    for i in command_list:
        ssh_con.send(i + "\r\n")
        time.sleep(1)
        output = ssh_con.recv(4096)
        print output
        if 'Error' in output:
            return False
        elif 'error' in output:
            return False
    return True


def telnet_command(IP, command_list, sheet_name):
    path = '/home/auto_config/' + IP + '.sh'
    
    if sheet_name == 'Juniper':
        create_cmd = 'cp /home/code/management/static/files/juniper.sh ' + path
        timeout = 3
    else:
        create_cmd = 'cp /home/code/management/static/files/dell-cisco.sh ' + path
        if sheet_name == 'Dell':
            if len(command_list) <= 10:
                timeout = 3
            elif len(command_list) <= 15:
                timeout = 4
            elif len(command_list) <= 20:
                timeout = 5
            elif len(command_list) <= 25:
                timeout = 6
            elif len(command_list) <= 30:
                timeout = 7
            elif len(command_list) <= 40:
                timeout = 9
            else:
                timeout = 10
        else:
            timeout = 4
    print create_cmd
    print timeout
    os.system(create_cmd)
    
    cmd = "sed -i s/---timeout---/" + str(timeout) + "/g " + path 
    print cmd
    os.system(cmd)
    
    cmd = "sed -i s/---IP---/" + IP + "/g " + path 
    print cmd
    os.system(cmd)
    
    
    command_list.reverse()
    
    if sheet_name == 'Juniper':
        command_cmd = "sed -i \'/\">\"/ a\send \"" + command_list[-1] + "\r\"\' " + path
        print command_cmd
        os.system(command_cmd)
        for command in command_list:
            if command == command_list[-1]:
                continue
            if '\n' in command:
                command = command.replace("\n","\\n")
            if '"' in command:
                command = command.replace('"','\\\\"')
    
            command_cmd = "sed -i \'/\"#\"/ a\send \"" + command + "\r\"\' " + path
            print command
            print command_cmd
            os.system(command_cmd)
    else:
        for command in command_list:
            if '\n' in command:
                command = command.replace("\n","\\n")
            if '"' in command:
                command = command.replace('"','\\\\"')
    
            command_cmd = "sed -i \'/\"#\"/ a\send \"" + command + "\r\"\' " + path
            print command
            print command_cmd
            os.system(command_cmd)
    
    output = commands.getoutput('sh ' + path)
    print output

    os.system('rm -rf ' + path)
    
    if 'error' in output:
        return False
    elif 'Error' in output:
        return False
    elif 'incorrect' in output:
        return False
    elif 'failed' in output:
        return False
    elif 'executing' in output:
        return False
    elif 'Invalid' in output:
        return False
    elif 'Connected' not in output:
        print 'aa'
        return False
    return True
        

